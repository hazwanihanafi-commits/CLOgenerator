# app.py
import os
import json
from io import BytesIO
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from openpyxl import Workbook, load_workbook

# -----------------------
# Basic setup
# -----------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, "static"),
    template_folder=os.path.join(BASE_DIR, "templates")
)

WORKBOOK_PATH = os.path.join(BASE_DIR, "SCLOG.xlsx")
FRONT_JSON = os.path.join(app.static_folder, "data", "SCLOG_front.json")

print("BOOT: STATIC =", app.static_folder)
print("BOOT: TEMPLATES =", app.template_folder)
print("BOOT: WORKBOOK =", WORKBOOK_PATH)
print("BOOT: FRONT_JSON =", FRONT_JSON)

# -----------------------
# Safe JSON loader
# -----------------------
def safe_load_json(path):
    if not os.path.exists(path):
        print("safe_load_json: file not found:", path)
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("safe_load_json: failed to parse json:", e)
        return {}

MAP = safe_load_json(FRONT_JSON)
# ensure keys exist
DEFAULT_KEYS = {
    "IEGs": [], "PEOs": [], "PLOs": [],
    "IEGtoPEO": {}, "PEOtoPLO": {},
    "PLOstatements": {}, "PEOstatements": {},
    "PLOtoVBE": {}, "PLOIndicators": {}, "SCmapping": {}
}
for k, v in DEFAULT_KEYS.items():
    MAP.setdefault(k, v)

# -----------------------
# Excel helpers (safe)
# -----------------------
def load_df(sheet_name):
    """Return pandas DataFrame for sheet_name or empty DataFrame on error."""
    if not os.path.exists(WORKBOOK_PATH):
        # workbook missing — return empty DF
        return pd.DataFrame()
    try:
        return pd.read_excel(WORKBOOK_PATH, sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        # sheet missing or read error
        print(f"load_df: cannot read sheet '{sheet_name}': {e}")
        return pd.DataFrame()

# PROFILE -> mapping sheet names
PROFILE_SHEET_MAP = {
    "health": "Mapping_health",
    "sc": "Mapping_sc",
    "eng": "Mapping_eng",
    "socs": "Mapping_socs",
    "edu": "Mapping_edu",
    "bus": "Mapping_bus",
    "arts": "Mapping_arts"
}

def get_mapping_sheet(profile):
    sheet = PROFILE_SHEET_MAP.get(profile, "Mapping")
    df = load_df(sheet)
    if df.empty:
        df = load_df("Mapping")
    return df

def get_plo_details(plo, profile="sc"):
    """
    Returns dict: SC_Code, SC_Desc, VBE, Domain
    or None if not found.
    """
    df = get_mapping_sheet(profile)
    if df.empty:
        return None
    # normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    col_plo = df.columns[0]
    mask = df[col_plo].astype(str).str.upper() == str(plo).upper()
    if not mask.any():
        return None
    row = df[mask].iloc[0]
    return {
        "SC_Code": row.get("SC Code", "") or row.get("SCCode", "") or "",
        "SC_Desc": row.get("SC Description", "") or row.get("SCDescription", "") or "",
        "VBE": row.get("VBE", "") or "",
        "Domain": row.get("Domain", "") or ""
    }

# -----------------------
# Meta (criterion & condition) from Excel "Criterion" sheet
# -----------------------
def get_meta_data(plo, bloom, profile="sc"):
    details = get_plo_details(plo, profile)
    if not details:
        return {}
    domain = (details.get("Domain") or "").lower()
    criterion = ""
    condition = ""
    df = load_df("Criterion")
    if not df.empty:
        df.columns = [c.strip() for c in df.columns]
        # safe matching
        left = df.iloc[:,0].astype(str).str.lower().fillna("")
        right = df.iloc[:,1].astype(str).str.lower().fillna("")
        mask = (left == domain) & (right == str(bloom).lower())
        if mask.any():
            row = df[mask].iloc[0]
            # cols 2 and 3 expected to be criterion and condition
            if len(row) > 2:
                criterion = str(row.iloc[2]) if row.iloc[2] is not None else ""
            if len(row) > 3:
                condition = str(row.iloc[3]) if row.iloc[3] is not None else ""
    if not condition:
        condition = {
            "cognitive": "interpreting tasks",
            "affective": "engaging with peers",
            "psychomotor": "performing skills"
        }.get(domain, "")
    connector = "by" if domain == "psychomotor" else "when"
    condition_final = f"{connector} {condition}"
    return {
        "sc_code": details.get("SC_Code",""),
        "sc_desc": details.get("SC_Desc",""),
        "vbe": details.get("VBE",""),
        "domain": domain,
        "criterion": criterion,
        "condition": condition_final
    }

# -----------------------
# Assessment & evidence engine
# -----------------------
def get_assessment(plo, bloom, domain):
    """
    Return a list of assessment methods (short strings) appropriate
    for the bloom & domain. If unknown, return empty list.
    """
    b = (bloom or "").lower().strip()
    d = (domain or "").lower().strip()
    cognitive = {
        "remember": ["MCQ", "Recall Quiz"],
        "understand": ["Short Answer Test", "Concept Explanation"],
        "apply": ["Case Study", "Problem-Solving Task"],
        "analyze": ["Data Analysis Task", "Critique Assignment"],
        "analyse": ["Data Analysis Task", "Critique Assignment"],
        "evaluate": ["Evaluation Report", "Evidence-Based Review"],
        "create": ["Design Project", "Research Proposal"]
    }
    psychomotor = {
        "perception": ["Observation Checklist", "Basic Skill Demonstration"],
        "set": ["Guided Task", "Preparation Checklist"],
        "guided response": ["Guided Skill Task", "Skills Test"],
        "mechanism": ["Skills Test", "OSCE"],
        "complex overt response": ["Integrated Practical", "OSCE"],
        "adaptation": ["Adapted Task Assessment", "Supervisor Eval"],
        "origination": ["Capstone Practical", "Innovation Deliverable"]
    }
    affective = {
        "receive": ["Reflection Log"],
        "respond": ["Participation Record", "Peer Review"],
        "value": ["Values Assignment", "Position Paper"],
        "organization": ["Group Portfolio"],
        "characterization": ["Professional Behaviour Assessment"]
    }
    if d == "psychomotor":
        return psychomotor.get(b, [])
    if d == "affective":
        return affective.get(b, [])
    return cognitive.get(b, [])

def get_evidence_for(assessment):
    a = assessment.lower()
    evidence_map = {
        "mcq": ["Score report", "Automated grading output"],
        "quiz": ["Quiz score", "Response pattern"],
        "short answer": ["Written answers", "Marker rubric"],
        "case study": ["Case analysis rubric", "Annotated report"],
        "problem-solving": ["Solution sheet", "Reasoning steps"],
        "critique": ["Critique essay", "Evaluator comments"],
        "evaluation report": ["Evaluation sheet", "Analytical justification"],
        "design project": ["Project files", "Prototype", "Design documentation"],
        "research proposal": ["Proposal draft", "Panel feedback"],
        "skill demonstration": ["Skills checklist", "Instructor feedback"],
        "osce": ["OSCE checklist", "Examiner score sheet"],
        "skills test": ["Performance score", "Competency checklist"],
        "clinical task": ["Clinical logbook", "Supervisor evaluation"],
        "reflection": ["Reflection journal", "Instructor comments"],
        "group": ["Group participation record", "Peer evaluation"],
        "portfolio": ["Portfolio files", "Growth documentation"],
    }
    for key in evidence_map:
        if key in a:
            return evidence_map[key]
    return ["Performance evidence", "Rubric score"]


# -----------------------
# Content suggestions (Option 1 fields)
# -----------------------
CONTENT_SUGGESTIONS = {
    "Computer Science": [
        "debug algorithms", "design software modules", "analyze data structures",
        "implement machine learning models", "develop APIs"
    ],
    "Medical & Health": [
        "interpret ECG waveforms", "assess patient vital signs", "perform clinical screenings",
        "analyze medical imaging", "evaluate rehabilitation progress"
    ],
    "Engineering": [
        "apply thermodynamics principles", "analyze structural loads",
        "simulate mechanical systems", "perform quality testing"
    ],
    "Social Sciences": [
        "evaluate community case studies", "analyze social policy impact",
        "interpret behavioral data", "conduct needs assessments"
    ],
    "Education": [
        "design learning activities", "evaluate student performance",
        "develop curriculum materials", "apply instructional strategies"
    ],
    "Business": [
        "analyze market trends", "evaluate financial reports",
        "develop business strategies", "conduct SWOT analysis"
    ],
    "Arts & Humanities": [
        "interpret visual artworks", "analyze literary texts",
        "evaluate cultural narratives", "produce creative concepts"
    ]
}

@app.route("/api/content/<field>")
def api_content(field):
    # field arrives url-encoded; try matching keys ignoring case
    key = None
    for k in CONTENT_SUGGESTIONS:
        if k.lower() == field.replace("%20"," ").lower():
            key = k
            break
    if not key:
        # attempt fuzzy partial match
        for k in CONTENT_SUGGESTIONS:
            if field.lower() in k.lower():
                key = k
                break
    return jsonify(CONTENT_SUGGESTIONS.get(key, []))

# -----------------------
# Mapping endpoints (IEG->PEO->PLO)
# -----------------------
@app.route("/api/mapping")
def api_mapping():
    return jsonify(MAP)

@app.route("/api/get_peos/<ieg>")
def api_get_peos(ieg):
    return jsonify(MAP.get("IEGtoPEO", {}).get(ieg, []))

@app.route("/api/get_plos/<peo>")
def api_get_plos(peo):
    return jsonify(MAP.get("PEOtoPLO", {}).get(peo, []))

@app.route("/api/get_statement/<level>/<stype>/<code>")
def api_get_statement(level, stype, code):
    stype = stype.upper()
    if stype == "PEO":
        return jsonify(MAP.get("PEOstatements", {}).get(level, {}).get(code, ""))
    if stype == "PLO":
        return jsonify(MAP.get("PLOstatements", {}).get(level, {}).get(code, ""))
    return jsonify("")

# -----------------------
# Bloom & Verb endpoints (from Excel)
# -----------------------
@app.route("/api/get_blooms/<plo>")
def api_get_blooms(plo):
    profile = request.args.get("profile", "sc").lower()
    details = get_plo_details(plo, profile)
    if not details:
        return jsonify([])
    domain = (details.get("Domain") or "").lower()
    sheet_map = {
        "cognitive": "Bloom_Cognitive",
        "affective": "Bloom_Affective",
        "psychomotor": "Bloom_Psychomotor"
    }
    sheet = sheet_map.get(domain, "Bloom_Cognitive")
    df = load_df(sheet)
    if df.empty:
        return jsonify([])
    blooms = df.iloc[:,0].dropna().astype(str).tolist()
    return jsonify(blooms)

@app.route("/api/get_verbs/<plo>/<bloom>")
def api_get_verbs(plo, bloom):
    profile = request.args.get("profile", "sc").lower()
    details = get_plo_details(plo, profile)
    if not details:
        return jsonify([])
    domain = (details.get("Domain") or "").lower()
    sheet_map = {
        "cognitive": "Bloom_Cognitive",
        "affective": "Bloom_Affective",
        "psychomotor": "Bloom_Psychomotor"
    }
    sheet = sheet_map.get(domain, "Bloom_Cognitive")
    df = load_df(sheet)
    if df.empty:
        return jsonify([])
    mask = df.iloc[:,0].astype(str).str.lower() == str(bloom).lower()
    if not mask.any():
        return jsonify([])
    raw = df[mask].iloc[0,1]
    verbs = [v.strip() for v in str(raw).split(",") if v.strip()]
    return jsonify(verbs)

# -----------------------
# Meta endpoint
# -----------------------
@app.route("/api/get_meta/<plo>/<bloom>")
def api_get_meta(plo, bloom):
    profile = request.args.get("profile", "sc").lower()
    return jsonify(get_meta_data(plo, bloom, profile))

# -----------------------
# Generate CLO
# -----------------------
# Simple global memory for last generated CLO (single-user)
LAST_CLO_DATA = {}

@app.route("/generate", methods=["POST"])
def generate():
    profile = request.form.get("profile", "sc").lower()

    plo = request.form.get("plo", "")
    bloom = request.form.get("bloom", "")
    verb = request.form.get("verb", "")
    content = request.form.get("content", "")
    level = request.form.get("level", "Degree")

    details = get_plo_details(plo, profile)
    if not details:
        return jsonify({"error": "Invalid PLO"}), 400

    domain = details["Domain"].lower()
    sc_desc = details["SC_Desc"]
    vbe = details["VBE"]

    # ------------------------------------------------------
    # CLEAN UP: Fix duplicated verb issue
    # ------------------------------------------------------
    content_words = content.strip().lower().split()

    # Common academic verbs list
    ACTION_VERBS = {
        "investigate","analyse","analyze","evaluate","interpret","assess","examine",
        "apply","perform","demonstrate","measure","design","explain"
    }

    # If content already starts with a verb → remove Bloom verb from content
    if content_words and content_words[0] in ACTION_VERBS:
        # Keep content starting ONE word after the verb
        content_clean = " ".join(content_words[1:])
    else:
        content_clean = content.strip()

    # ------------------------------------------------------
    # META extraction
    # ------------------------------------------------------
    meta_res = get_meta_data(plo, bloom, profile)

    condition_core = (
        meta_res["condition"]
        .replace("when ", "")
        .replace("by ", "")
        .strip()
    )

    criterion = meta_res["criterion"]
    connector = "when" if domain != "psychomotor" else "by"

    # ------------------------------------------------------
    # CLO generation (cleaned)
    # ------------------------------------------------------
    clo = (
        f"{verb.lower()} {content_clean} using {sc_desc.lower()} "
        f"{connector} {condition_core} guided by {vbe.lower()}."
    ).capitalize()

    # ------------------------------------------------------
    # Variant generation
    # ------------------------------------------------------
    variants = {
        "Standard": clo,
        "Critical Thinking": clo.replace("using", "critically using"),
        "Action": clo.replace("when", "while"),
    }

    # ------------------------------------------------------
    # IEG → PEO mapping
    # ------------------------------------------------------
    peo = None
    ieg = None

    for p, plos in MAP["PEOtoPLO"].items():
        if plo in plos:
            peo = p
            break

    for i, peos in MAP["IEGtoPEO"].items():
        if peo in peos:
            ieg = i
            break

    # ------------------------------------------------------
    # Statements
    # ------------------------------------------------------
    plo_statement = MAP["PLOstatements"][level].get(plo, "")
    peo_statement = MAP["PEOstatements"][level].get(peo, "")
    plo_indicator = MAP["PLOIndicators"].get(plo, "")

    # ------------------------------------------------------
    # Assessment & Evidence suggestions
    # ------------------------------------------------------
    assessments = get_assessment(plo, bloom, domain)

    evidence_output = {}
    for a in assessments:
        evidence_output[a] = get_evidence_for(a)

    # ------------------------------------------------------
    # Save for Excel download (now includes variants)
    # ------------------------------------------------------
    global LAST_CLO_DATA
    LAST_CLO_DATA = {
        "clo": clo,
        "variants": variants,           # ← NEW: store all variant CLOs
        "plo": plo,
        "peo": peo,
        "ieg": ieg,
        "plo_statement": plo_statement,
        "peo_statement": peo_statement,
        "plo_indicator": plo_indicator,
        "sc_code": details["SC_Code"],
        "sc_desc": sc_desc,
        "vbe": vbe,
        "domain": domain,
        "criterion": criterion,
        "condition": condition_core,
        "assessments": assessments,
        "evidence": evidence_output,
        "rubric": {
            "indicator": f"Ability to {verb.lower()} {sc_desc.lower()}",
            "excellent": "Performs at an excellent level",
            "good": "Performs well",
            "satisfactory": "Meets minimum level",
            "poor": "Below expected"
        }
    }

    return jsonify({
        "clo": clo,
        "clo_options": variants,
        "peo": peo,
        "ieg": ieg,
        "sc_code": details["SC_Code"],
        "sc_desc": sc_desc,
        "vbe": vbe,
        "criterion": criterion,
        "condition": condition_core,
        "plo_statement": plo_statement,
        "peo_statement": peo_statement,
        "plo_indicator": plo_indicator,
        "assessments": assessments,
        "evidence": evidence_output,
        "plo_indicator": MAP.get("PLOIndicators", {}).get(plo, "")

    })

# -----------------------
# Downloads
# -----------------------
@app.route("/download")
def download_clo():
    global LAST_CLO_DATA
    if not LAST_CLO_DATA:
        return "No CLO available. Generate one first.", 400

    data = LAST_CLO_DATA
    wb = Workbook()
    ws = wb.active
    ws.title = "CLO"

    ws.append(["Field","Value"])
    ws.append(["CLO", data.get("clo","")])
    ws.append(["PLO", data.get("plo","")])
    ws.append(["PLO statement", data.get("plo_statement","")])
    ws.append(["PEO", data.get("peo","")])
    ws.append(["PEO statement", data.get("peo_statement","")])
    ws.append(["PLO indicator", data.get("plo_indicator","")])
    ws.append(["SC code", data.get("sc_code","")])
    ws.append(["SC description", data.get("sc_desc","")])
    ws.append(["VBE", data.get("vbe","")])
    ws.append(["Domain", data.get("domain","")])
    ws.append(["Criterion", data.get("criterion","")])
    ws.append(["Condition", data.get("condition","")])

    # assessments + evidence
    ws.append([])
    ws.append(["Assessment method","Suggested evidence"])
    for a in data.get("assessments", []):
        ev = "; ".join(data.get("evidence", {}).get(a, []))
        ws.append([a, ev])

    # rubric
    ws.append([])
    ws.append(["Rubric Indicator", data["rubric"]["indicator"]])
    ws.append(["Excellent", data["rubric"]["excellent"]])
    ws.append(["Good", data["rubric"]["good"]])
    ws.append(["Satisfactory", data["rubric"]["satisfactory"]])
    ws.append(["Poor", data["rubric"]["poor"]])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"CLO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download_rubric")
def download_rubric():
    global LAST_CLO_DATA
    if not LAST_CLO_DATA:
        return "No rubric available. Generate first.", 400
    data = LAST_CLO_DATA
    wb = Workbook()
    ws = wb.active
    ws.title = "Rubric"
    ws.append(["Rubric Component","Description"])
    ws.append(["Indicator", data["rubric"]["indicator"]])
    ws.append(["Excellent", data["rubric"]["excellent"]])
    ws.append(["Good", data["rubric"]["good"]])
    ws.append(["Satisfactory", data["rubric"]["satisfactory"]])
    ws.append(["Poor", data["rubric"]["poor"]])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"Rubric_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -----------------------
# UI route
# -----------------------
@app.route("/")
def index():
    # generator.html uses the mapping loaded via /api/mapping
    return render_template("generator.html")
# ================================================
# IEG → PEO LOGIC EXPLANATIONS
# ================================================
@app.route("/api/logic/ieg_peo/<ieg>", methods=["GET"])
def logic_ieg_peo(ieg):
    logic_map = {
        "IEG1": (
            "IEG1 focuses on knowledge, critical thinking, and problem-solving. "
            "PEO1 is mapped because it operationalizes analytical and intellectual competencies "
            "into graduate outcomes aligned with the discipline."
        ),
        "IEG2": (
            "IEG2 emphasises altruistic values, ethics, professionalism, and scientific thinking. "
            "PEO2 supports this by shaping graduates with integrity, responsibility and citizenship."
        ),
        "IEG3": (
            "IEG3 promotes socio-entrepreneurial mindset and societal wellbeing. "
            "PEO3 is mapped because it directs graduates toward contributing to sustainability and equity."
        ),
        "IEG4": (
            "IEG4 strengthens communication competence across disciplines, society and technology. "
            "PEO4 aligns because it emphasises effective communication and interaction in academic and industry settings."
        ),
        "IEG5": (
            "IEG5 emphasises leadership, teamwork and lifelong learning. "
            "PEO5 is mapped because it equips graduates with collaboration and continuous development skills."
        ),
    }

    return logic_map.get(ieg, "No logic explanation available."), 200


# ================================================
# PEO → PLO LOGIC EXPLANATIONS
# (Nice to have, more complete experience)
# ================================================
@app.route("/api/logic/peo_plo/<peo>/<plo>", methods=["GET"])
def logic_peo_plo(peo, plo):
    logic_map = {
        "PEO1": {
            "PLO1": "PLO1 supports PEO1 by ensuring students acquire fundamental disciplinary knowledge.",
            "PLO2": "PLO2 supports PEO1 by reinforcing critical thinking and problem-solving skills.",
            "PLO3": "PLO3 strengthens analytical and practical abilities required by PEO1.",
            "PLO6": "PLO6 ensures students can apply knowledge in real-world contexts, supporting PEO1.",
            "PLO7": "PLO7 aligns with PEO1 through exposure to complex challenges requiring critical reasoning."
        },
        "PEO2": {
            "PLO11": "PLO11 emphasises ethics, professionalism and responsibility, directly supporting PEO2."
        },
        "PEO3": {
            "PLO10": "PLO10 supports PEO3 by preparing students to address societal and sustainable development issues.",
            "PLO9": "PLO9 contributes to PEO3 by fostering awareness of global, societal and sustainability challenges."
        },
        "PEO4": {
            "PLO5": "PLO5 aligns strongly with PEO4 by focusing on effective communication skills."
        },
        "PEO5": {
            "PLO4": "PLO4 supports PEO5 through teamwork and collaborative competencies.",
            "PLO8": "PLO8 reinforces leadership, adaptability and lifelong learning required for PEO5.",
            "PLO9": "PLO9 also supports PEO5 by encouraging engagement with complex real-world problems."
        }
    }

    # Return explanation if both exist
    if peo in logic_map and plo in logic_map[peo]:
        return logic_map[peo][plo], 200

    return "No PEO → PLO logic available.", 200

from flask import Flask, jsonify, request, send_file, make_response
from io import BytesIO
import json

app = Flask(__name__)

# ---------------------------
# Demo mapping data (example)
# ---------------------------
MAP = {
    "IEGs": ["IEG1", "IEG2", "IEG3", "IEG4", "IEG5"],
    "PEOs": ["PEO1","PEO2","PEO3","PEO4","PEO5"],
    "PLOs": ["PLO1","PLO2","PLO3","PLO4","PLO5","PLO6","PLO7","PLO8","PLO9","PLO10","PLO11"],
    "PLOIndicators": {
        "PLO1":"Indicator A","PLO2":"Indicator B","PLO3":"Indicator C",
        "PLO4":"Indicator D","PLO5":"Indicator E","PLO6":"Indicator F",
        "PLO7":"Indicator G","PLO8":"Indicator H","PLO9":"Indicator I",
        "PLO10":"Indicator J","PLO11":"Indicator K"
    }
}

PEO_TO_PLO = {
    "PEO1": ["PLO1","PLO2","PLO3","PLO6","PLO7"],
    "PEO2": ["PLO11"],
    "PEO3": ["PLO10","PLO9"],
    "PEO4": ["PLO5"],
    "PEO5": ["PLO8","PLO4","PLO9"]
}

IEG_TO_PEO = {
    "IEG1": ["PEO1"],
    "IEG2": ["PEO2"],
    "IEG3": ["PEO3"],
    "IEG4": ["PEO4"],
    "IEG5": ["PEO5"]
}

# ---------------------------
# Simple verbs/blooms metadata (demo)
# ---------------------------
PLO_BLOOMS = {
    "PLO1": ["Remember","Understand","Apply"],
    "PLO2": ["Understand","Analyze"],
    "PLO3": ["Apply","Analyze","Evaluate"],
    "PLO4": ["Apply","Create"],
    "PLO5": ["Remember","Understand","Apply"],
    "PLO6": ["Apply"],
    "PLO7": ["Analyze"],
    "PLO8": ["Create","Evaluate"],
    "PLO9": ["Analyze","Evaluate"],
    "PLO10": ["Create"],
    "PLO11": ["Evaluate"]
}

VERBS = {
    ("PLO1","Remember"): ["list","define","recall"],
    ("PLO2","Analyze"): ["differentiate","compare","contrast"],
    ("PLO3","Apply"): ["demonstrate","use","implement"],
    ("PLO4","Create"): ["design","construct","formulate"],
    ("PLO5","Apply"): ["perform","execute"]
}

META_SAMPLE = {
    "sc_code": "SC-101",
    "sc_desc": "Sample Skill/Competency description",
    "vbe": "Verbs/Behaviors/Examples",
    "domain": "Cognitive",
    "condition": "Given a dataset",
    "criterion": "80% accuracy"
}

# ---------------------------
# Logic explanation routes
# ---------------------------
@app.route("/api/logic/ieg_peo/<ieg>", methods=["GET"])
def logic_ieg_peo(ieg):
    logic_map = {
        "IEG1": "IEG1 focuses on knowledge, critical thinking and problem-solving. PEO1 is mapped because it operationalizes these competencies into program-level outcomes.",
        "IEG2": "IEG2 emphasises altruistic values, ethics and professionalism. PEO2 supports building graduates with strong character and civic responsibility.",
        "IEG3": "IEG3 promotes socio-entrepreneurial skills and sustainability; PEO3 encourages application of skills to societal wellbeing.",
        "IEG4": "IEG4 stresses effective communication; PEO4 aligns by developing communication and confidence.",
        "IEG5": "IEG5 highlights leadership, teamwork and lifelong learning; PEO5 focuses on fostering these transferable skills."
    }
    return logic_map.get(ieg, "No logic explanation available."), 200, {'Content-Type':'text/plain; charset=utf-8'}


@app.route("/api/logic/peo_plo/<peo>/<plo>", methods=["GET"])
def logic_peo_plo(peo, plo):
    logic_map = {
        "PEO1": {
            "PLO1": "PLO1 supports PEO1 by ensuring students acquire fundamental disciplinary knowledge.",
            "PLO2": "PLO2 supports PEO1 by reinforcing critical thinking and problem solving.",
            "PLO3": "PLO3 strengthens analytical and practical abilities.",
            "PLO6": "PLO6 ensures students apply knowledge in practical contexts.",
            "PLO7": "PLO7 exposes students to complex problems requiring higher-order thought."
        },
        "PEO2": {"PLO11": "PLO11 emphasizes ethics and professionalism, thus supporting PEO2."},
        "PEO3": {"PLO10": "PLO10 prepares students to design solutions for societal problems.", "PLO9":"PLO9 focuses on sustainability and social impact."},
        "PEO4": {"PLO5":"PLO5 aims to improve communication skills, aligned with PEO4."},
        "PEO5": {"PLO4":"PLO4 builds teamwork and leadership skills.","PLO8":"PLO8 develops leadership & lifelong learning competencies.","PLO9":"PLO9 encourages engagement with complex issues."}
    }
    if peo in logic_map and plo in logic_map[peo]:
        return logic_map[peo][plo], 200, {'Content-Type':'text/plain; charset=utf-8'}
    return "No PEO → PLO logic available.", 200, {'Content-Type':'text/plain; charset=utf-8'}

# ---------------------------
# Mapping endpoints used by frontend
# ---------------------------
@app.route("/api/mapping")
def api_mapping():
    return jsonify(MAP)

@app.route("/api/get_peos/<ieg>")
def api_get_peos(ieg):
    # Return PEOs mapped to IEG (demo)
    return jsonify(IEG_TO_PEO.get(ieg, []))

@app.route("/api/get_plos/<peo>")
def api_get_plos(peo):
    return jsonify(PEO_TO_PLO.get(peo, []))

@app.route("/api/get_blooms/<plo>")
def api_get_blooms(plo):
    return jsonify(PLO_BLOOMS.get(plo, ["Remember","Understand","Apply"]))

@app.route("/api/get_verbs/<plo>/<bloom>")
def api_get_verbs(plo, bloom):
    key = (plo, bloom)
    return jsonify(VERBS.get(key, ["identify","explain"]))

@app.route("/api/get_meta/<plo>/<bloom>")
def api_get_meta(plo, bloom):
    # Return sample metadata; replace with your real lookup
    return jsonify(META_SAMPLE)

@app.route("/api/get_statement/<level>/<typ>/<code>")
def api_get_statement(level, typ, code):
    # typ can be PEO or PLO - demo return
    if typ == "PEO":
        return jsonify(f"PEO statement for {code} at level {level}.")
    return jsonify(f"PLO statement for {code} at level {level}.")

@app.route("/api/content/<field>")
def api_content(field):
    examples = {
        "Computer Science": ["implement sort algorithms","design REST APIs","write unit tests"],
        "Medical & Health": ["interpret ECG waveforms","perform basic life support","measure blood pressure"],
        "Engineering": ["analyze stress-strain curves","design a cantilever beam","model thermodynamic cycles"],
        "Education": ["design lesson plans","apply formative assessment","classroom management strategies"],
        "Business": ["analyze financial statements","design marketing strategies","evaluate business models"],
        "Social Sciences": ["conduct a survey","apply qualitative coding","interpret statistical output"],
        "Arts & Humanities": ["critique a painting","analyze a poem","contextualize a historical event"]
    }
    return jsonify(examples.get(field, []))

# ---------------------------
# Generate API - demo logic
# ---------------------------
@app.route("/generate", methods=["POST"])
def generate_clo():
    profile = request.form.get('profile', 'sc')
    level = request.form.get('level', 'Degree')
    plo = request.form.get('plo', '')
    bloom = request.form.get('bloom', '')
    verb = request.form.get('verb', '')
    content = request.form.get('content', '')
    # Simple templated CLO as demo
    clo_text = f"Students will be able to {verb} {content} ({bloom}) — mapped to {plo}."
    data = {
        "clo": clo_text,
        "clo_options": {
            "Formal": clo_text,
            "Simplified": f"{verb.capitalize()} {content}."
        },
        "assessments": ["Written exam","Practical test"],
        "evidence": {"Assignment": ["report", "presentation"]},
        "peo_statement": f"PEO statement demo for mapping {plo}.",
        "plo_statement": f"PLO statement demo for {plo}.",
        "plo_indicator": MAP["PLOIndicators"].get(plo, "")
    }
    return jsonify(data)

# ---------------------------
# Download demo endpoints
# ---------------------------
@app.route("/download")
def download_clo():
    # demo text file
    content = "Generated CLO (demo)\n\nThis is a demo download. Replace with your generated file implementation."
    return make_text_download(content, "generated_clo.txt")

@app.route("/download_rubric")
def download_rubric():
    content = "Generated Rubric (demo)\n\nThis is a demo rubric file."
    return make_text_download(content, "rubric.txt")

def make_text_download(text, filename):
    buffer = BytesIO()
    buffer.write(text.encode('utf-8'))
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='text/plain')

# ---------------------------
# Run server (for local testing)
# ---------------------------
if __name__ == "__main__":
    app.run(debug=True, port=5000)


# -----------------------
# Run
# -----------------------
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")




